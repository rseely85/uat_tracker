import io
import os
from datetime import datetime

from flask import Flask, jsonify, render_template, request, send_file

from init_db import DB_PATH, get_connection, init_db

app = Flask(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def ok(data):
    return jsonify({"success": True, "data": data})


def err(msg, status=400):
    return jsonify({"success": False, "error": msg}), status


def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def rows_to_list(rows):
    return [dict(r) for r in rows]


def resolve_labels(conn, case):
    """Add human-readable label for each d1..d7 id on a test case dict."""
    for i in range(1, 8):
        key = f"d{i}_id"
        val = case.get(key)
        if val:
            r = conn.execute(
                "SELECT value FROM dropdown_items WHERE item_id = ?", (val,)
            ).fetchone()
            case[f"d{i}_label"] = r["value"] if r else ""
        else:
            case[f"d{i}_label"] = ""
    return case


# ---------------------------------------------------------------------------
# Main page
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Dropdowns
# ---------------------------------------------------------------------------

@app.route("/api/dropdowns", methods=["GET"])
def get_all_dropdowns():
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM dropdown_items ORDER BY dropdown_num, sort_order, item_id"
    ).fetchall()
    conn.close()
    return ok(rows_to_list(rows))


@app.route("/api/dropdowns/<int:num>", methods=["GET"])
def get_dropdown_by_num(num):
    parent_id = request.args.get("parent_id", type=int)
    conn = get_connection()
    if parent_id is not None:
        rows = conn.execute(
            "SELECT * FROM dropdown_items WHERE dropdown_num=? AND parent_item_id=? ORDER BY sort_order, item_id",
            (num, parent_id),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM dropdown_items WHERE dropdown_num=? ORDER BY sort_order, item_id",
            (num,),
        ).fetchall()
    conn.close()
    return ok(rows_to_list(rows))


@app.route("/api/dropdowns", methods=["POST"])
def add_dropdown_item():
    data = request.get_json(force=True) or {}
    dropdown_num = data.get("dropdown_num")
    value = data.get("value", "").strip()
    parent_item_id = data.get("parent_item_id")
    sort_order = data.get("sort_order", 0)

    if not dropdown_num or not value:
        return err("dropdown_num and value are required")

    conn = get_connection()
    c = conn.cursor()
    c.execute(
        "INSERT INTO dropdown_items (dropdown_num, value, parent_item_id, sort_order) VALUES (?,?,?,?)",
        (dropdown_num, value, parent_item_id, sort_order),
    )
    new_id = c.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM dropdown_items WHERE item_id=?", (new_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row)), 201


@app.route("/api/dropdowns/<int:item_id>", methods=["PUT"])
def update_dropdown_item(item_id):
    data = request.get_json(force=True) or {}
    conn = get_connection()
    item = conn.execute("SELECT * FROM dropdown_items WHERE item_id=?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return err("Item not found", 404)

    value = data.get("value", item["value"])
    sort_order = data.get("sort_order", item["sort_order"])
    parent_item_id = data.get("parent_item_id", item["parent_item_id"])

    conn.execute(
        "UPDATE dropdown_items SET value=?, sort_order=?, parent_item_id=? WHERE item_id=?",
        (value, sort_order, parent_item_id, item_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM dropdown_items WHERE item_id=?", (item_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row))


@app.route("/api/dropdowns/<int:item_id>", methods=["DELETE"])
def delete_dropdown_item(item_id):
    try:
        conn = get_connection()
        conn.execute("DELETE FROM dropdown_items WHERE item_id=?", (item_id,))
        conn.commit()
        conn.close()
        return ok({"deleted": item_id})
    except Exception as e:
        return err(str(e), 500)


# ---------------------------------------------------------------------------
# Test Cases
# ---------------------------------------------------------------------------

@app.route("/api/test-cases", methods=["GET"])
def list_test_cases():
    filters = []
    params = []
    for i in range(1, 8):
        val = request.args.get(f"d{i}", type=int)
        if val:
            filters.append(f"tc.d{i}_id = ?")
            params.append(val)

    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    conn = get_connection()
    rows = conn.execute(
        f"""SELECT tc.* FROM test_cases tc
            LEFT JOIN dropdown_items d3 ON d3.item_id = tc.d3_id
            LEFT JOIN dropdown_items d4 ON d4.item_id = tc.d4_id
            LEFT JOIN dropdown_items d5 ON d5.item_id = tc.d5_id
            LEFT JOIN dropdown_items d6 ON d6.item_id = tc.d6_id
            {where}
            ORDER BY COALESCE(d3.sort_order, 999),
                     COALESCE(d4.sort_order, 999),
                     COALESCE(d5.sort_order, 999),
                     COALESCE(d6.sort_order, 999),
                     tc.case_id""",
        params
    ).fetchall()
    cases = [resolve_labels(conn, dict(r)) for r in rows]
    conn.close()
    return ok(cases)


@app.route("/api/test-cases", methods=["POST"])
def create_test_case():
    data = request.get_json(force=True) or {}
    conn = get_connection()
    c = conn.cursor()
    c.execute(
        """INSERT INTO test_cases
           (d1_id, d2_id, d3_id, d4_id, d5_id, d6_id, d7_id, task, notes)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (
            data.get("d1_id"),
            data.get("d2_id"),
            data.get("d3_id"),
            data.get("d4_id"),
            data.get("d5_id"),
            data.get("d6_id"),
            data.get("d7_id"),
            data.get("task", ""),
            data.get("notes", ""),
        ),
    )
    new_id = c.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM test_cases WHERE case_id=?", (new_id,)).fetchone()
    case = resolve_labels(conn, dict(row))
    conn.close()
    return ok(case), 201


@app.route("/api/test-cases/<int:case_id>", methods=["GET"])
def get_test_case(case_id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM test_cases WHERE case_id=?", (case_id,)).fetchone()
    if not row:
        conn.close()
        return err("Not found", 404)
    case = resolve_labels(conn, dict(row))
    conn.close()
    return ok(case)


@app.route("/api/test-cases/<int:case_id>", methods=["PUT"])
def update_test_case(case_id):
    data = request.get_json(force=True) or {}
    conn = get_connection()
    row = conn.execute("SELECT * FROM test_cases WHERE case_id=?", (case_id,)).fetchone()
    if not row:
        conn.close()
        return err("Not found", 404)

    old = dict(row)
    conn.execute(
        """UPDATE test_cases SET
           d1_id=?, d2_id=?, d3_id=?, d4_id=?, d5_id=?, d6_id=?, d7_id=?,
           task=?, notes=?, updated_at=datetime('now')
           WHERE case_id=?""",
        (
            data.get("d1_id", old["d1_id"]),
            data.get("d2_id", old["d2_id"]),
            data.get("d3_id", old["d3_id"]),
            data.get("d4_id", old["d4_id"]),
            data.get("d5_id", old["d5_id"]),
            data.get("d6_id", old["d6_id"]),
            data.get("d7_id", old["d7_id"]),
            data.get("task", old["task"]),
            data.get("notes", old["notes"]),
            case_id,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM test_cases WHERE case_id=?", (case_id,)).fetchone()
    case = resolve_labels(conn, dict(row))
    conn.close()
    return ok(case)


@app.route("/api/test-cases/<int:case_id>", methods=["DELETE"])
def delete_test_case(case_id):
    try:
        conn = get_connection()
        # Remove dependent rows first (test_results has no CASCADE on case_id)
        conn.execute("DELETE FROM test_results WHERE case_id=?", (case_id,))
        conn.execute("DELETE FROM test_cases WHERE case_id=?", (case_id,))
        conn.commit()
        conn.close()
        return ok({"deleted": case_id})
    except Exception as e:
        return err(str(e), 500)


# ---------------------------------------------------------------------------
# Test Groups
# ---------------------------------------------------------------------------

@app.route("/api/test-groups", methods=["GET"])
def list_test_groups():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM test_groups ORDER BY created_at DESC").fetchall()
    conn.close()
    return ok(rows_to_list(rows))


@app.route("/api/test-groups", methods=["POST"])
def create_test_group():
    data = request.get_json(force=True) or {}
    name = data.get("group_name", "").strip()
    if not name:
        return err("group_name is required")
    conn = get_connection()
    c = conn.cursor()
    c.execute(
        "INSERT INTO test_groups (group_name, description) VALUES (?,?)",
        (name, data.get("description", "")),
    )
    new_id = c.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM test_groups WHERE group_id=?", (new_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row)), 201


@app.route("/api/test-groups/<int:group_id>", methods=["GET"])
def get_test_group(group_id):
    conn = get_connection()
    group = conn.execute("SELECT * FROM test_groups WHERE group_id=?", (group_id,)).fetchone()
    if not group:
        conn.close()
        return err("Not found", 404)
    cases_rows = conn.execute(
        """SELECT tc.*, tgc.sort_order AS group_sort
           FROM test_cases tc
           JOIN test_group_cases tgc ON tgc.case_id = tc.case_id
           WHERE tgc.group_id = ?
           ORDER BY tgc.sort_order, tc.case_id""",
        (group_id,),
    ).fetchall()
    cases = [resolve_labels(conn, dict(r)) for r in cases_rows]
    result = dict(group)
    result["cases"] = cases
    conn.close()
    return ok(result)


@app.route("/api/test-groups/<int:group_id>", methods=["PUT"])
def update_test_group(group_id):
    data = request.get_json(force=True) or {}
    conn = get_connection()
    row = conn.execute("SELECT * FROM test_groups WHERE group_id=?", (group_id,)).fetchone()
    if not row:
        conn.close()
        return err("Not found", 404)
    old = dict(row)
    conn.execute(
        "UPDATE test_groups SET group_name=?, description=? WHERE group_id=?",
        (data.get("group_name", old["group_name"]), data.get("description", old["description"]), group_id),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM test_groups WHERE group_id=?", (group_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row))


@app.route("/api/test-groups/<int:group_id>", methods=["DELETE"])
def delete_test_group(group_id):
    try:
        conn = get_connection()
        conn.execute("DELETE FROM test_groups WHERE group_id=?", (group_id,))
        conn.commit()
        conn.close()
        return ok({"deleted": group_id})
    except Exception as e:
        return err(str(e), 500)


@app.route("/api/test-groups/<int:group_id>/cases", methods=["POST"])
def add_case_to_group(group_id):
    data = request.get_json(force=True) or {}
    case_id = data.get("case_id")
    sort_order = data.get("sort_order", 0)
    if not case_id:
        return err("case_id is required")
    conn = get_connection()
    # Avoid duplicates
    existing = conn.execute(
        "SELECT id FROM test_group_cases WHERE group_id=? AND case_id=?", (group_id, case_id)
    ).fetchone()
    if existing:
        conn.close()
        return err("Case already in group")
    conn.execute(
        "INSERT INTO test_group_cases (group_id, case_id, sort_order) VALUES (?,?,?)",
        (group_id, case_id, sort_order),
    )
    conn.commit()
    conn.close()
    return ok({"group_id": group_id, "case_id": case_id}), 201


@app.route("/api/test-groups/<int:group_id>/cases/<int:case_id>", methods=["DELETE"])
def remove_case_from_group(group_id, case_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM test_group_cases WHERE group_id=? AND case_id=?", (group_id, case_id)
    )
    conn.commit()
    conn.close()
    return ok({"removed": case_id})


# ---------------------------------------------------------------------------
# Test Runs
# ---------------------------------------------------------------------------

@app.route("/api/test-runs", methods=["GET"])
def list_test_runs():
    conn = get_connection()
    rows = conn.execute(
        """SELECT tr.*,
              tg.group_name,
              COUNT(res.result_id) AS total,
              SUM(CASE WHEN res.result='pass' THEN 1 ELSE 0 END) AS pass_count,
              SUM(CASE WHEN res.result='fail' THEN 1 ELSE 0 END) AS fail_count,
              SUM(CASE WHEN res.result='skip' THEN 1 ELSE 0 END) AS skip_count,
              SUM(CASE WHEN res.result='pending' THEN 1 ELSE 0 END) AS pending_count
           FROM test_runs tr
           LEFT JOIN test_groups tg ON tg.group_id = tr.group_id
           LEFT JOIN test_results res ON res.run_id = tr.run_id
           GROUP BY tr.run_id
           ORDER BY tr.run_date DESC""",
    ).fetchall()
    conn.close()
    return ok(rows_to_list(rows))


@app.route("/api/test-runs", methods=["POST"])
def create_test_run():
    data = request.get_json(force=True) or {}
    group_id = data.get("group_id")
    run_name = data.get("run_name", "")
    notes = data.get("notes", "")

    if not group_id:
        return err("group_id is required")

    conn = get_connection()
    group = conn.execute("SELECT * FROM test_groups WHERE group_id=?", (group_id,)).fetchone()
    if not group:
        conn.close()
        return err("Group not found", 404)

    if not run_name:
        today = datetime.now().strftime("%Y-%m-%d")
        run_name = f"{group['group_name']} — {today}"

    c = conn.cursor()
    c.execute(
        "INSERT INTO test_runs (group_id, run_name, notes) VALUES (?,?,?)",
        (group_id, run_name, notes),
    )
    run_id = c.lastrowid

    # Auto-create pending results for all cases in group
    cases = conn.execute(
        "SELECT case_id FROM test_group_cases WHERE group_id=? ORDER BY sort_order, case_id",
        (group_id,),
    ).fetchall()
    for row in cases:
        c.execute(
            "INSERT INTO test_results (run_id, case_id, result) VALUES (?,?,'pending')",
            (run_id, row["case_id"]),
        )

    conn.commit()
    run = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(run)), 201


@app.route("/api/test-runs/<int:run_id>", methods=["GET"])
def get_test_run(run_id):
    conn = get_connection()
    run = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not run:
        conn.close()
        return err("Not found", 404)

    results = conn.execute(
        """SELECT res.*, tc.d1_id, tc.d2_id, tc.d3_id, tc.d4_id,
                  tc.d5_id, tc.d6_id, tc.d7_id, tc.task, tc.notes AS case_notes
           FROM test_results res
           JOIN test_cases tc ON tc.case_id = res.case_id
           LEFT JOIN dropdown_items d3 ON d3.item_id = tc.d3_id
           LEFT JOIN dropdown_items d4 ON d4.item_id = tc.d4_id
           LEFT JOIN dropdown_items d5 ON d5.item_id = tc.d5_id
           LEFT JOIN dropdown_items d6 ON d6.item_id = tc.d6_id
           WHERE res.run_id = ?
           ORDER BY COALESCE(d3.sort_order, 999),
                    COALESCE(d4.sort_order, 999),
                    COALESCE(d5.sort_order, 999),
                    COALESCE(d6.sort_order, 999),
                    res.result_id""",
        (run_id,),
    ).fetchall()

    resolved = [resolve_labels(conn, dict(r)) for r in results]
    result_obj = dict(run)
    result_obj["results"] = resolved
    conn.close()
    return ok(result_obj)


@app.route("/api/test-runs/<int:run_id>", methods=["PUT"])
def update_test_run(run_id):
    data = request.get_json(force=True) or {}
    conn = get_connection()
    run = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not run:
        conn.close()
        return err("Not found", 404)
    old = dict(run)
    conn.execute(
        "UPDATE test_runs SET run_name=?, notes=?, status=? WHERE run_id=?",
        (
            data.get("run_name", old["run_name"]),
            data.get("notes", old["notes"]),
            data.get("status", old["status"]),
            run_id,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row))


@app.route("/api/test-runs/<int:run_id>/results/<int:case_id>", methods=["POST"])
def save_result(run_id, case_id):
    data = request.get_json(force=True) or {}
    result = data.get("result", "pending")
    result_notes = data.get("result_notes", "")
    completed_at = datetime.now().isoformat() if result != "pending" else None

    conn = get_connection()
    existing = conn.execute(
        "SELECT result_id FROM test_results WHERE run_id=? AND case_id=?", (run_id, case_id)
    ).fetchone()

    if existing:
        conn.execute(
            """UPDATE test_results SET result=?, result_notes=?, completed_at=?
               WHERE run_id=? AND case_id=?""",
            (result, result_notes, completed_at, run_id, case_id),
        )
    else:
        conn.execute(
            """INSERT INTO test_results (run_id, case_id, result, result_notes, completed_at)
               VALUES (?,?,?,?,?)""",
            (run_id, case_id, result, result_notes, completed_at),
        )

    conn.commit()
    row = conn.execute(
        "SELECT * FROM test_results WHERE run_id=? AND case_id=?", (run_id, case_id)
    ).fetchone()
    conn.close()
    return ok(row_to_dict(row))


@app.route("/api/test-runs/<int:run_id>/complete", methods=["PUT"])
def complete_test_run(run_id):
    conn = get_connection()
    conn.execute("UPDATE test_runs SET status='complete' WHERE run_id=?", (run_id,))
    conn.commit()
    row = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    conn.close()
    return ok(row_to_dict(row))


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def _build_run_sheet(ws, run_id, conn):
    """Populate an openpyxl worksheet with results for a single run."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    headers = [
        "Run Name", "Run Date", "Group",
        "App", "Testing Group", "Menu Item", "Section",
        "Field/Button", "Function 1", "Function 2",
        "Task", "Result", "Result Notes", "Completed At",
    ]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    run = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (run_id,)).fetchone()
    if not run:
        return

    group_name = ""
    if run["group_id"]:
        g = conn.execute("SELECT group_name FROM test_groups WHERE group_id=?", (run["group_id"],)).fetchone()
        group_name = g["group_name"] if g else ""

    results = conn.execute(
        """SELECT res.*, tc.d1_id, tc.d2_id, tc.d3_id, tc.d4_id,
                  tc.d5_id, tc.d6_id, tc.d7_id, tc.task
           FROM test_results res
           JOIN test_cases tc ON tc.case_id = res.case_id
           WHERE res.run_id = ?
           ORDER BY res.result_id""",
        (run_id,),
    ).fetchall()

    def label(item_id):
        if not item_id:
            return ""
        r = conn.execute("SELECT value FROM dropdown_items WHERE item_id=?", (item_id,)).fetchone()
        return r["value"] if r else ""

    for r in results:
        ws.append([
            run["run_name"],
            run["run_date"],
            group_name,
            label(r["d1_id"]),
            label(r["d2_id"]),
            label(r["d3_id"]),
            label(r["d4_id"]),
            label(r["d5_id"]),
            label(r["d6_id"]),
            label(r["d7_id"]),
            r["task"],
            r["result"],
            r["result_notes"] or "",
            r["completed_at"] or "",
        ])

    # Auto-width
    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


@app.route("/api/export/excel/<run_id>", methods=["GET"])
def export_run_excel(run_id):
    import openpyxl
    conn = get_connection()

    if run_id == "all":
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        runs = conn.execute("SELECT run_id, run_name FROM test_runs ORDER BY run_date DESC").fetchall()
        for run in runs:
            safe_name = str(run["run_name"] or f"Run {run['run_id']}")[:31]
            ws = wb.create_sheet(title=safe_name)
            _build_run_sheet(ws, run["run_id"], conn)
        filename = "all_runs.xlsx"
    else:
        try:
            rid = int(run_id)
        except ValueError:
            conn.close()
            return err("Invalid run_id")
        run = conn.execute("SELECT * FROM test_runs WHERE run_id=?", (rid,)).fetchone()
        if not run:
            conn.close()
            return err("Not found", 404)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = str(run["run_name"] or f"Run {rid}")[:31]
        _build_run_sheet(ws, rid, conn)
        filename = f"run_{rid}.xlsx"

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# Keep the old /all route working as an alias
@app.route("/api/export/excel/all", methods=["GET"])
def export_all_excel():
    return export_run_excel("all")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    init_db()
    print("UAT Tracker running at http://127.0.0.1:5050")
    app.run(debug=True, port=5050)
